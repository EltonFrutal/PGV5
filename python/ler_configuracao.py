# ============================================================
#  ler_configuracao.py — Lê Configuracao.xlsx e gera config.py
# ============================================================

import openpyxl
import os

def ler_excel_config(caminho_excel):
    """Lê arquivo Configuracao.xlsx e retorna dicionários."""
    wb = openpyxl.load_workbook(caminho_excel)
    
    # Ler USUARIOS
    ws_user = wb['usuario']
    usuarios = []
    for row in ws_user.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Se tem ID
            usuarios.append({
                'id': row[0],
                'email': row[1],
                'senha': str(row[2]),
                'perfil': row[3],
                'permissao': row[4],
                'foto': row[5] if len(row) > 5 else None
            })
    
    # Ler EMPRESAS
    ws_emp = wb['empresa']
    empresas = []
    for row in ws_emp.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[2]).lower() == 'sim':  # Apenas empresas ativas
            empresas.append({
                'id': row[0],
                'nome': row[1],
                'ativo': row[2],
                'arquivo': row[3],
                'logo': row[4] if len(row) > 4 else None
            })
    
    # Ler USUARIOEMPRESA (permissões)
    ws_useremp = wb['usuarioempresa']
    usuario_empresas = []
    for row in ws_useremp.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Se tem usuario_id
            usuario_empresas.append({
                'usuario_id': row[0],
                'organizacao_id': row[2],  # Coluna organizacao
                'exibe_vendas': str(row[4]).lower() == 'sim',
                'exibe_compras': str(row[5]).lower() == 'sim',
                'exibe_areceber': str(row[6]).lower() == 'sim',
                'exibe_apagar': str(row[7]).lower() == 'sim',
                'exibe_historicos': str(row[8]).lower() == 'sim',
                'exibe_indicadores': str(row[9]).lower() == 'sim',
                'exibe_resultados': str(row[10]).lower() == 'sim',
            })
    
    return usuarios, empresas, usuario_empresas


def gerar_config_py(usuarios, empresas, usuario_empresas, caminho_base_dados):
    """Gera conteúdo do config.py baseado nos dados do Excel."""
    
    # Mapeamento de IDs para nomes
    emp_map = {e['id']: e['nome'] for e in empresas}
    
    # Cores e ícones por empresa
    cores_icones = {
        'Multicar': {'cor': '#10b981', 'icone': '🏪'},
        'RS': {'cor': '#6366f1', 'icone': '🚗'},
        'Default': {'cor': '#8b5cf6', 'icone': '🏢'}
    }
    
    # Gerar lista de USUARIOS
    usuarios_str = "USUARIOS = [\n"
    for u in usuarios:
        # Buscar empresas deste usuário
        empresas_usuario = [
            emp_map[ue['organizacao_id']] 
            for ue in usuario_empresas 
            if ue['usuario_id'] == u['id'] and ue['organizacao_id'] in emp_map
        ]
        
        foto_arquivo = f"Imagem/fotos/{u['foto']}.jpg" if u['foto'] else None
        
        usuarios_str += f"""    {{
        "email":    "{u['email']}",
        "senha":    "{u['senha']}",
        "nome":     "{u['email'].split('@')[0].title()}",
        "perfil":   "{u['perfil'].lower()}",
        "empresas": {empresas_usuario},
        "foto":     "{foto_arquivo}" if foto_arquivo else "null",
    }},
"""
    usuarios_str += "]\n"
    
    # Gerar lista de EMPRESAS
    empresas_str = "EMPRESAS = [\n"
    for e in empresas:
        nome = e['nome']
        config = cores_icones.get(nome, cores_icones['Default'])
        logo_arquivo = f"Imagem/Logos/{e['logo']}.png" if e['logo'] else None
        
        empresas_str += f"""    {{
        "id":               "{nome.replace(' ', '_')}",
        "nome":             "{nome}",
        "arquivo":          "{nome.lower()}.html",
        "cor":              "{config['cor']}",
        "icone":            "{config['icone']}",
        "logo":             "{logo_arquivo}" if logo_arquivo else "null",
        "base":             r"{caminho_base_dados}\\{e['arquivo']}",
        "vendas_arquivo":   "{nome.lower()}_vendas_dash.html",
        "listagem_arquivo": "{nome.lower()}_vendas_lista.html",
    }},
"""
    empresas_str += "]\n"
    
    # Template completo do config.py
    config_template = f'''# ============================================================
#  config.py — Configurações do PGV5
#  GERADO AUTOMATICAMENTE de Configuracao.xlsx
#  NÃO EDITE MANUALMENTE - Execute: python ler_configuracao.py
# ============================================================

# ── REPOSITÓRIO ──────────────────────────────────────────────
REPO_PATH = r"G:\\Meu Drive\\Consys\\PainelGerencial\\PGV5"

# ── USUÁRIOS ─────────────────────────────────────────────────
# senha: texto puro — será criptografado automaticamente
{usuarios_str}

# ── EMPRESAS ─────────────────────────────────────────────────
{empresas_str}

# ── MÓDULOS ──────────────────────────────────────────────────
# pronto: True = ativo, False = "Em breve"
MODULOS = [
    {{"id":"vendas",      "label":"Vendas",      "desc":"Faturamento, pedidos e vendedores","cor":"#6366f1","pronto":True}},
    {{"id":"compras",     "label":"Compras",     "desc":"Entradas e fornecedores",          "cor":"#10b981","pronto":False}},
    {{"id":"a_receber",   "label":"A Receber",   "desc":"Títulos e inadimplência",          "cor":"#f59e0b","pronto":False}},
    {{"id":"a_pagar",     "label":"A Pagar",     "desc":"Obrigações e vencimentos",         "cor":"#ef4444","pronto":False}},
    {{"id":"historicos",  "label":"Históricos",  "desc":"Movimentos por período",           "cor":"#8b5cf6","pronto":False}},
    {{"id":"indicadores", "label":"Indicadores", "desc":"KPIs e metas",                     "cor":"#06b6d4","pronto":False}},
    {{"id":"resultados",  "label":"Resultados",  "desc":"DRE e resultado final",            "cor":"#f97316","pronto":False}},
]

# ── FONTES ───────────────────────────────────────────────────
FONT_URL    = "https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700;800&display=swap"
FONT_FAMILY = "'DM Sans', sans-serif"

FONT_URL_LISTA    = "https://fonts.googleapis.com/css2?family=Nunito:wght@400;500;600;700;800;900&display=swap"
FONT_FAMILY_LISTA = "'Nunito', sans-serif"
'''
    
    return config_template


def main():
    """Script principal."""
    print("="*60)
    print("  Gerando config.py a partir de Configuracao.xlsx")
    print("="*60)
    
    # Caminhos
    excel_path = r"G:\Meu Drive\Consys\PainelGerencial\Configuracao.xlsx"
    config_path = r"G:\Meu Drive\Consys\PainelGerencial\PGV5\Python\config.py"
    base_dados_path = r"G:\Meu Drive\Consys\PainelGerencial\BaseDados"
    
    print(f"\n[1/3] Lendo {excel_path}...")
    usuarios, empresas, usuario_empresas = ler_excel_config(excel_path)
    
    print(f"      ✔ {len(usuarios)} usuários")
    print(f"      ✔ {len(empresas)} empresas ativas")
    print(f"      ✔ {len(usuario_empresas)} permissões")
    
    print(f"\n[2/3] Gerando config.py...")
    config_content = gerar_config_py(usuarios, empresas, usuario_empresas, base_dados_path)
    
    print(f"\n[3/3] Salvando em {config_path}...")
    with open(config_path, 'w', encoding='utf-8') as f:
        f.write(config_content)
    
    print("      ✔ config.py atualizado!")
    print("\n" + "="*60)
    print("  Pronto! Execute: python gerar_pgv5.py")
    print("="*60)


if __name__ == "__main__":
    main()
